"""
update-xlsx.py — called by offer-scraper.js to update price cells in place.
Usage: python3 update-xlsx.py <file> <sheet> <url_col> <rarity_col> <lowest_col> <highest_col> <average_col> <count_col> <json_stats>

All column indices are 1-based (matching Excel convention).
json_stats is a JSON array of {offer_url, rarity, lowest, highest, average, offer_count}.
"""

import sys
import json

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "ERROR: openpyxl is not installed.\n"
        "Install it with one of:\n"
        "  Arch Linux:  sudo pacman -S python-openpyxl\n"
        "  pip (other): pip install openpyxl\n"
        "  pip (Arch):  pip install openpyxl --break-system-packages"
    )
    sys.exit(2)

def main():
    if len(sys.argv) < 10:
        print("Usage: update-xlsx.py <file> <sheet> <url_col> <rarity_col> <lowest_col> <highest_col> <average_col> <count_col> <json_stats>")
        sys.exit(1)

    file_path  = sys.argv[1]
    sheet_name = sys.argv[2]
    url_col    = int(sys.argv[3])
    rarity_col = int(sys.argv[4])
    lowest_col  = int(sys.argv[5])
    highest_col = int(sys.argv[6])
    average_col = int(sys.argv[7])
    count_col   = int(sys.argv[8])
    stats       = json.loads(sys.argv[9])

    # Build lookup: "url::rarity" -> stat
    lookup = {}
    for s in stats:
        key = f"{s['offer_url']}::{s['rarity']}"
        lookup[key] = s

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    updated = 0
    for row in ws.iter_rows(min_row=2):
        url    = str(row[url_col - 1].value    or "").strip()
        rarity = str(row[rarity_col - 1].value or "").strip()
        key    = f"{url}::{rarity}"
        stat   = lookup.get(key)
        if not stat:
            continue

        row[lowest_col  - 1].value = round(stat["lowest"],   2)
        row[highest_col - 1].value = round(stat["highest"],  2)
        row[average_col - 1].value = round(stat["average"],  2)
        row[count_col   - 1].value = stat["offer_count"]
        updated += 1

    wb.save(file_path)
    print(f"updated:{updated}")

main()
